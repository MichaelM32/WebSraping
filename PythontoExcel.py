import openpyxl as xl 
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')



ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size=24, italic=False, bold=True)

myfont = Font(name='Times New Roman', size=24, italic=False, bold=True)


ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Allignment'

ws.merge_cells('A1:B1')

ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] =  150

ws['A8'] = 'Total'
ws['A8'].font = myfont

ws['B8'] = '=SUM(B2:B4)'


ws.column_dimensions['A'].width = 25

#wb.save('PythonToExcel.xlsx')

#####
##Read the excel file = 'ProduceReport.xlsx that you created earlier
##Write all the contents of this file to 'second sheet' in the current workbook

##display the grand total and average of 'amt sold' and 'total'
##at the bottom of the list along with appropriate labels

write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row


for row in read_ws.iter_rows(min_row=1, max_row= maxR, max_col= maxC):
    for cell in row:
        write_sheet[cell.coordinate] = cell.value


wb.save('PythonToExcel.xlsx')